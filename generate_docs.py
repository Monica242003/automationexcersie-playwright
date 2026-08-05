import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Create docs directory if it doesn't exist
os.makedirs("docs", exist_ok=True)

test_cases = [
    {
        "id": "TC-01",
        "name": "Register User",
        "description": "Verify that a new user can successfully register, view the 'Logged in as' header, delete the account, and verify the 'ACCOUNT DELETED!' screen.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Signup / Login' button",
            "Verify 'New User Signup!' is visible",
            "Enter name and email address",
            "Click 'Signup' button",
            "Verify 'ENTER ACCOUNT INFORMATION' is visible",
            "Fill details: Title, Name, Password, Date of birth",
            "Select checkbox 'Sign up for our newsletter!'",
            "Select checkbox 'Receive special offers from our partners!'",
            "Fill address details: First name, Last name, Company, Address, Address2, Country, State, City, Zipcode, Mobile Number",
            "Click 'Create Account' button",
            "Verify 'ACCOUNT CREATED!' is visible",
            "Click 'Continue' button",
            "Verify 'Logged in as John David' is visible",
            "Click 'Delete Account' button",
            "Verify 'ACCOUNT DELETED!' is visible",
            "Click 'Continue' button"
        ],
        "expected_result": "User is successfully registered, logged in, and the account is deleted successfully."
    },
    {
        "id": "TC-02",
        "name": "Login User with correct email and password",
        "description": "Verify that a registered user can log in with a correct email and password, delete the account, and verify deletion.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Signup / Login' button",
            "Verify 'Login to your account' is visible",
            "Register a temporary user 'LoginCorrect' with email and password",
            "Click on 'Signup / Login' button",
            "Enter correct email and password",
            "Click 'login' button",
            "Verify 'Logged in as LoginCorrect' is visible",
            "Click 'Delete Account' button",
            "Verify 'ACCOUNT DELETED!' is visible"
        ],
        "expected_result": "User logs in successfully, user name is displayed, and user account is deleted."
    },
    {
        "id": "TC-03",
        "name": "Login User with incorrect email and password",
        "description": "Verify that a user cannot log in with an incorrect email and password, and receives a validation error message.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Signup / Login' button",
            "Verify 'Login to your account' is visible",
            "Enter incorrect email and password",
            "Click 'login' button",
            "Verify error 'Your email or password is incorrect!' is visible"
        ],
        "expected_result": "Login fails and the error message 'Your email or password is incorrect!' is displayed."
    },
    {
        "id": "TC-04",
        "name": "Logout User",
        "description": "Verify that a logged-in user can log out successfully and is navigated back to the login page.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Signup / Login' button",
            "Verify 'Login to your account' is visible",
            "Register a temporary user and log in",
            "Verify 'Logged in as LogoutUser' is visible",
            "Click 'Logout' button",
            "Verify user is navigated to login page"
        ],
        "expected_result": "User is logged out successfully and redirected to the login/signup page."
    },
    {
        "id": "TC-05",
        "name": "Register User with existing email",
        "description": "Verify that a user cannot register with an email address that is already registered, and receives a validation error message.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Signup / Login' button",
            "Verify 'New User Signup!' is visible",
            "Register a temporary user first",
            "Enter name and existing email address, and click 'Signup' button",
            "Verify error 'Email Address already exist!' is visible"
        ],
        "expected_result": "Registration fails and the error message 'Email Address already exist!' is displayed."
    },
    {
        "id": "TC-09",
        "name": "Search Product",
        "description": "Verify that a user can search for a specific product and see the search results matching the query.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Products' button",
            "Verify user is navigated to ALL PRODUCTS page successfully",
            "Enter product name 'Blue Top' in search input and click search button",
            "Verify 'SEARCHED PRODUCTS' is visible",
            "Verify all the products related to search are visible"
        ],
        "expected_result": "The search results display all products related to 'Blue Top' successfully."
    },
    {
        "id": "TC-12",
        "name": "Add Products in Cart",
        "description": "Verify that a user can add multiple products to the cart and verify their prices, quantity, and total price.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click 'Products' button",
            "Hover over first product and click 'Add to cart'",
            "Click 'Continue Shopping' button",
            "Hover over second product and click 'Add to cart'",
            "Click 'View Cart' button",
            "Verify both products are added to Cart",
            "Verify their prices, quantity, and total price"
        ],
        "expected_result": "Both products are added to the cart, displaying correct price, quantity, and total price."
    },
    {
        "id": "TC-14",
        "name": "Place Order: Register while Checkout",
        "description": "Verify that a user can add a product to the cart, proceed to checkout, register a new account, complete the order with payment, and delete the account.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click 'Products' button",
            "Hover over first product and click 'Add to cart'",
            "Click 'View Cart' button",
            "Verify cart page is displayed",
            "Click Proceed To Checkout",
            "Click 'Register / Login' button",
            "Fill details in Signup and create account",
            "Verify 'ACCOUNT CREATED!' is visible and click 'Continue'",
            "Verify 'Logged in as CheckoutRegister' is visible",
            "Click 'Cart' button, click Proceed To Checkout",
            "Verify Address Details and Review Your Order",
            "Enter description in comment text area and click 'Place Order'",
            "Enter payment details: Name on Card, Card Number, CVC, Expiration Date",
            "Click 'Pay and Confirm Order' button",
            "Verify success message 'Your order has been placed successfully!'",
            "Click 'Delete Account' button",
            "Verify 'ACCOUNT DELETED!' is visible and click 'Continue'"
        ],
        "expected_result": "Order is placed successfully, confirmation message is shown, and the registered user account is deleted."
    },
    {
        "id": "TC-21",
        "name": "Add review on product",
        "description": "Verify that a user can write and submit a review for a product, and see a success confirmation message.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Click on 'Products' button",
            "Verify user is navigated to ALL PRODUCTS page successfully",
            "Click on 'View Product' of first product",
            "Verify 'Write Your Review' is visible",
            "Enter name, email, and review comment",
            "Click 'Submit' review button",
            "Verify review success message 'Thank you for your review.' is visible"
        ],
        "expected_result": "Product review is submitted successfully and confirmation message is displayed."
    },
    {
        "id": "TC-23",
        "name": "Verify address details in checkout page",
        "description": "Verify that delivery and billing addresses on the checkout page match the addresses entered during account registration.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click on 'Signup / Login' button",
            "Register a new account 'AddressVerify'",
            "Verify 'ACCOUNT CREATED!' and click 'Continue'",
            "Verify 'Logged in as AddressVerify' is visible",
            "Click 'Products' button, hover over first product and click 'Add to cart'",
            "Click 'View Cart' button and verify cart page is displayed",
            "Click Proceed To Checkout",
            "Verify delivery address is the same as registration address",
            "Verify billing address is the same as registration address",
            "Click 'Delete Account' button and verify deletion"
        ],
        "expected_result": "Delivery and billing address details on checkout match registration address details."
    },
    {
        "id": "TC-24",
        "name": "Download Invoice after purchase order",
        "description": "Verify that a user can purchase an order, download the invoice, and verify that the invoice is downloaded successfully.",
        "steps": [
            "Navigate to URL 'http://automationexercise.com'",
            "Verify home page is visible successfully",
            "Click 'Products' button, hover over first product and click 'Add to cart'",
            "Click 'View Cart' button and verify cart page is displayed",
            "Click Proceed To Checkout",
            "Click 'Register / Login' button and register a new account",
            "Verify 'ACCOUNT CREATED!' is visible and click 'Continue'",
            "Verify 'Logged in as InvoiceDownload' is visible",
            "Click 'Cart' button and click Proceed To Checkout",
            "Verify Address Details and Review Your Order",
            "Enter comment description and click 'Place Order'",
            "Enter payment details, click 'Pay and Confirm Order'",
            "Verify success message 'Your order has been placed successfully!'",
            "Click 'Download Invoice' button and verify invoice is downloaded successfully",
            "Click 'Continue' button",
            "Click 'Delete Account' button, verify deletion, and click 'Continue'"
        ],
        "expected_result": "Order is placed successfully, invoice file is downloaded, and the account is deleted."
    }
]

def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    
    regular_font = Font(name=font_family, size=10)
    tc_id_font = Font(name=font_family, size=10, bold=True)
    
    border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Sheet Title
    ws.cell(row=2, column=2, value="Automation Exercise - Test Suite").font = title_font
    
    # Table Headers
    headers = ["Test Case ID", "Test Case Name", "Description", "Steps", "Expected Result"]
    header_row = 4
    
    for col_num, header in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Fill Data
    current_row = header_row + 1
    for tc in test_cases:
        # TC ID
        c_id = ws.cell(row=current_row, column=2, value=tc["id"])
        c_id.font = tc_id_font
        c_id.alignment = Alignment(horizontal="center", vertical="top")
        c_id.border = thin_border
        
        # Name
        c_name = ws.cell(row=current_row, column=3, value=tc["name"])
        c_name.font = regular_font
        c_name.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c_name.border = thin_border
        
        # Description
        c_desc = ws.cell(row=current_row, column=4, value=tc["description"])
        c_desc.font = regular_font
        c_desc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c_desc.border = thin_border
        
        # Steps
        steps_text = "\n".join([f"{i+1}. {step}" for i, step in enumerate(tc["steps"])])
        c_steps = ws.cell(row=current_row, column=5, value=steps_text)
        c_steps.font = regular_font
        c_steps.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c_steps.border = thin_border
        
        # Expected Result
        c_exp = ws.cell(row=current_row, column=6, value=tc["expected_result"])
        c_exp.font = regular_font
        c_exp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c_exp.border = thin_border
        
        current_row += 1
        
    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15  # TC ID
    ws.column_dimensions['C'].width = 35  # Name
    ws.column_dimensions['D'].width = 45  # Description
    ws.column_dimensions['E'].width = 60  # Steps
    ws.column_dimensions['F'].width = 40  # Expected Result
    
    # Row Heights
    ws.row_dimensions[2].height = 25  # Title row
    ws.row_dimensions[3].height = 10  # Blank spacer row
    ws.row_dimensions[4].height = 28  # Header row
    for r in range(5, current_row):
        ws.row_dimensions[r].height = 120 # General padding height for multi-line steps
        
    file_path = os.path.join("docs", "Automation_Exercise_Test_Cases.xlsx")
    wb.save(file_path)
    print(f"Excel file successfully generated at: {file_path}")

class TestSuitePDF(FPDF):
    def header(self):
        # Blue top accent bar
        self.set_fill_color(31, 78, 120) # 1F4E78
        self.rect(0, 0, 210, 8, "F")
        
        self.set_y(12)
        self.set_font("helvetica", "B", 10)
        self.set_text_color(100, 100, 100)
        self.cell(0, 10, "Automation Exercise - Test Suite Specification", border=0, align="L")
        
        # Right aligned page number in header
        self.set_font("helvetica", "I", 9)
        # Shift X position to align to right
        self.set_x(180)
        self.cell(20, 10, f"Page {self.page_no()}", border=0, align="R")
        
        # Light grey divider line
        self.set_draw_color(220, 220, 220)
        self.line(10, 22, 200, 22)
        self.ln(5)

    def footer(self):
        # Divider line
        self.set_draw_color(220, 220, 220)
        self.line(10, 282, 200, 282)
        
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(120, 10, "Confidential - E2E Playwright Cucumber Automation Suite", border=0, align="L")
        self.set_x(170)
        self.cell(30, 10, "Generated: July 2026", border=0, align="R")

def generate_pdf():
    pdf = TestSuitePDF()
    pdf.set_margins(10, 15, 10)
    pdf.add_page()
    
    # Document Title
    pdf.set_y(25)
    pdf.set_font("helvetica", "B", 20)
    pdf.set_text_color(31, 78, 120)
    pdf.cell(0, 12, "E2E Test Suite Specification", align="L")
    pdf.ln(12)
    
    pdf.set_font("helvetica", "", 10)
    pdf.set_text_color(80, 80, 80)
    pdf.multi_cell(0, 5, "This document lists the official End-to-End automation test cases for the Automation Exercise website. "
                         "All scenarios defined here are automated using Playwright, TypeScript, and Cucumber BDD features.", 0, "L")
    pdf.ln(5)
    
    # Table of contents/summary table
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(31, 78, 120)
    pdf.cell(0, 8, "Summary Table of Test Cases", align="L")
    pdf.ln(8)
    
    # Table headers
    pdf.set_font("helvetica", "B", 10)
    pdf.set_text_color(255, 255, 255)
    pdf.set_fill_color(31, 78, 120)
    pdf.cell(20, 8, "ID", border=1, align="C", fill=True)
    pdf.cell(125, 8, "Test Case Name", border=1, align="L", fill=True)
    pdf.cell(45, 8, "Status", border=1, align="C", fill=True)
    pdf.ln(8)
    
    # Table contents
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(50, 50, 50)
    for tc in test_cases:
        pdf.cell(20, 7, tc["id"], border=1, align="C")
        pdf.cell(125, 7, tc["name"], border=1, align="L")
        pdf.set_text_color(40, 120, 40)
        pdf.set_font("helvetica", "B", 9)
        pdf.cell(45, 7, "AUTOMATED", border=1, align="C")
        pdf.ln(7)
        pdf.set_text_color(50, 50, 50)
        pdf.set_font("helvetica", "", 9)
        
    pdf.ln(8)
    
    # Detailed Test Cases
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(31, 78, 120)
    pdf.cell(0, 10, "Detailed Test Case Specifications", align="L")
    pdf.ln(10)
    
    for tc in test_cases:
        # Check height before starting a new block to avoid orphaned titles at the bottom of the page
        if pdf.get_y() > 210:
            pdf.add_page()
            
        # Test Case Title Header
        pdf.set_font("helvetica", "B", 11)
        pdf.set_text_color(31, 78, 120)
        pdf.set_fill_color(230, 240, 250)
        pdf.cell(0, 8, f" {tc['id']}: {tc['name']}", border=0, align="L", fill=True)
        pdf.ln(8)
        
        # Description
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(25, 5, "Description:", border=0, align="L")
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(50, 50, 50)
        pdf.multi_cell(0, 5, tc["description"], 0, "L")
        pdf.ln(2)
        
        # Steps
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(25, 5, "Test Steps:", border=0, align="L")
        pdf.ln(5)
        
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(50, 50, 50)
        for i, step in enumerate(tc["steps"]):
            pdf.set_x(20)
            pdf.set_font("helvetica", "B", 9)
            pdf.cell(8, 5, f"{i+1}.", border=0, align="L")
            pdf.set_font("helvetica", "", 9)
            pdf.multi_cell(172, 5, step, 0, "L")
        pdf.ln(2)
            
        # Expected Result
        pdf.set_x(10)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(27, 5, "Expected Result:", border=0, align="L")
        pdf.set_font("helvetica", "I", 9)
        pdf.set_text_color(31, 78, 120)
        pdf.multi_cell(0, 5, tc["expected_result"], 0, "L")
        
        pdf.ln(6)
        # Accent separator line
        pdf.set_draw_color(240, 240, 240)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(4)
        
    file_path = os.path.join("docs", "Automation_Exercise_Test_Cases.pdf")
    pdf.output(file_path)
    print(f"PDF file successfully generated at: {file_path}")

if __name__ == "__main__":
    generate_excel()
    generate_pdf()
